"""
Face_Mask_Detection.py  ─  Enhanced Crowd Detection with Recording & Analytics
===============================================================================
Controls:
  R    ─  Start / Stop recording session
  ESC  ─  Quit

While recording:
  • Every detected face per frame is logged to CSV + Excel report
  • Live on-screen dashboard: FPS, per-label counts, rolling trend graph
  • Session summary printed to console on stop

Output files (auto-named with timestamp, saved in sessions/ folder):
  sessions/session_YYYYMMDD_HHMMSS.csv
  sessions/session_YYYYMMDD_HHMMSS.xlsx
"""

import cv2
import numpy as np
from ultralytics import YOLO
from keras.models import load_model
from collections import deque
import time
import csv
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("⚠  openpyxl not installed — CSV only.  pip install openpyxl")

# ══════════════════════════════════════════════════════════════
#  CONFIGURATION
# ══════════════════════════════════════════════════════════════
YOLO_CONF_THRESHOLD = 0.50
MASK_CONF_THRESHOLD = 0.65
SMOOTHING_WINDOW    = 7
FACE_PADDING_RATIO  = 0.15
INPUT_SIZE          = (224, 224)
GRAPH_WINDOW_SEC    = 30      # seconds shown in rolling trend graph
GRAPH_H             = 110     # height of graph area in pixels
OUTPUT_DIR          = "sessions"

# ══════════════════════════════════════════════════════════════
#  LOAD MODELS
# ══════════════════════════════════════════════════════════════
yolo_model = YOLO("yolov8n-face.pt")
mask_model = load_model("best_model.keras")

# ══════════════════════════════════════════════════════════════
#  LABELS & COLOURS  (BGR)
# ══════════════════════════════════════════════════════════════
LABELS = ["Incorrect Mask", "Correct Mask", "No Mask"]

BBOX_COLOR = {
    "Correct Mask":   (0,   200,   0),
    "No Mask":        (0,   0,   220),
    "Incorrect Mask": (0,   200, 255),
    "Unknown":        (180, 180, 180),
}

# Dashboard colours (BGR)
D_BG        = (18,  18,  28)
D_TEXT      = (210, 210, 210)
D_ACCENT    = (0,   210, 255)
D_CORRECT   = (0,   200,   0)
D_NOMASK    = (0,   0,   220)
D_INCORRECT = (0,   200, 255)
D_UNKNOWN   = (120, 120, 120)

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ══════════════════════════════════════════════════════════════
#  SMOOTHER & TRACKER
# ══════════════════════════════════════════════════════════════
class FaceSmoother:
    def __init__(self, window=SMOOTHING_WINDOW):
        self.history = deque(maxlen=window)

    def update(self, probs):
        self.history.append(probs)
        avg  = np.mean(self.history, axis=0)
        cid  = int(np.argmax(avg))
        conf = float(avg[cid])
        if conf < MASK_CONF_THRESHOLD:
            return "Unknown", conf
        return LABELS[cid], conf


class FaceTracker:
    def __init__(self, max_dist=80, window=SMOOTHING_WINDOW):
        self.smoothers  = {}
        self.centroids  = {}
        self.max_dist   = max_dist
        self.window     = window
        self._nid       = 0

    def _cen(self, x1, y1, x2, y2):
        return ((x1+x2)//2, (y1+y2)//2)

    def update(self, dets):
        if not dets:
            self.smoothers.clear(); self.centroids.clear()
            return []
        new_cens    = [self._cen(*d[:4]) for d in dets]
        matched     = {}
        for i, c in enumerate(new_cens):
            best_id, best_d = None, self.max_dist
            for tid, pc in self.centroids.items():
                d = np.hypot(c[0]-pc[0], c[1]-pc[1])
                if d < best_d and tid not in matched.values():
                    best_d, best_id = d, tid
            if best_id is None:
                best_id = self._nid
                self.smoothers[best_id] = FaceSmoother(self.window)
                self._nid += 1
            matched[i] = best_id
        active = set(matched.values())
        for tid in list(self.smoothers):
            if tid not in active:
                del self.smoothers[tid]; del self.centroids[tid]
        res = []
        for i, (x1, y1, x2, y2, probs) in enumerate(dets):
            tid = matched[i]
            label, conf = self.smoothers[tid].update(probs)
            self.centroids[tid] = new_cens[i]
            res.append((x1, y1, x2, y2, label, conf))
        return res


# ══════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════
def pad_bbox(x1, y1, x2, y2, r, fh, fw):
    pw = int((x2-x1)*r); ph = int((y2-y1)*r)
    return max(0,x1-pw), max(0,y1-ph), min(fw,x2+pw), min(fh,y2+ph)

def preprocess(bgr):
    rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
    rgb = cv2.resize(rgb, INPUT_SIZE).astype("float32") / 255.0
    return np.expand_dims(rgb, 0)

def draw_label(frame, text, x1, y1, color):
    font = cv2.FONT_HERSHEY_SIMPLEX
    (tw, th), bl = cv2.getTextSize(text, font, 0.60, 2)
    p = 5
    ry1 = max(0, y1 - th - 2*p - bl)
    ry2 = max(th+bl, y1)
    cv2.rectangle(frame, (x1, ry1), (x1+tw+2*p, ry2), color, -1)
    cv2.putText(frame, text, (x1+p, ry2-bl-p), font, 0.60, (255,255,255), 2, cv2.LINE_AA)


# ══════════════════════════════════════════════════════════════
#  CSV / EXCEL
# ══════════════════════════════════════════════════════════════
CSV_HDR = ["Frame","Timestamp","Elapsed_s","Face_ID","Label",
           "Confidence_%","Total_Faces","Correct_Mask","No_Mask",
           "Incorrect_Mask","FPS"]

def init_csv(path):
    with open(path, "w", newline="") as f:
        csv.writer(f).writerow(CSV_HDR)

def append_csv(path, rows):
    with open(path, "a", newline="") as f:
        w = csv.writer(f)
        for r in rows: w.writerow(r)

def export_excel(csv_path, xlsx_path, sess_start, sess_end, summary):
    if not EXCEL_OK:
        return
    import csv as _csv
    wb  = openpyxl.Workbook()

    # ── Detection Log sheet ───────────────────────────────────
    ws  = wb.active
    ws.title = "Detection Log"
    thin   = Side(style="thin", color="CCCCCC")
    bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="12122A")
    h_font = Font(bold=True, color="00D2FF", name="Consolas", size=10)
    a_fill = PatternFill("solid", fgColor="F2F5FF")
    label_fill = {
        "Correct Mask":   "D4F7D4",
        "No Mask":        "F7D4D4",
        "Incorrect Mask": "FFF5D4",
        "Unknown":        "E8E8E8",
    }

    with open(csv_path, newline="") as f:
        for ri, row in enumerate(_csv.reader(f), 1):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(ri, ci, val)
                cell.border    = bdr
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if ri == 1:
                    cell.font = h_font; cell.fill = h_fill
                else:
                    if ri % 2 == 0: cell.fill = a_fill
                    if ci == 5 and val in label_fill:
                        cell.fill = PatternFill("solid", fgColor=label_fill[val])
                        cell.font = Font(bold=True, name="Consolas", size=9)

    for i, w in enumerate([8,22,11,9,16,15,13,14,10,16,7], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    # ── Summary sheet ────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    tf  = Font(bold=True, size=15, color="12122A", name="Consolas")
    lf  = Font(bold=True, size=11, color="333333", name="Consolas")
    vf  = Font(size=11,            color="12122A", name="Consolas")
    af  = PatternFill("solid", fgColor="EEF2FF")

    ws2["B2"] = "FACE MASK DETECTION — SESSION REPORT"
    ws2["B2"].font = tf

    rows = [
        ("Session Start",          sess_start),
        ("Session End",            sess_end),
        ("Duration",               summary["duration"]),
        ("Frames Logged",          summary["total_frames"]),
        ("Total Face Detections",  summary["total_detections"]),
        ("✅  Correct Mask",       summary["correct"]),
        ("❌  No Mask",            summary["no_mask"]),
        ("⚠   Incorrect Mask",    summary["incorrect"]),
        ("❓  Unknown",            summary["unknown"]),
        ("Peak FPS",               f"{summary['peak_fps']:.1f}"),
        ("Average FPS",            f"{summary['avg_fps']:.1f}"),
    ]
    for i, (lbl, val) in enumerate(rows, 4):
        lc = ws2.cell(i, 2, lbl); lc.font = lf
        vc = ws2.cell(i, 4, str(val)); vc.font = vf
        if i % 2 == 0:
            for c in range(2, 6):
                ws2.cell(i, c).fill = af

    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["D"].width = 28

    wb.save(xlsx_path)
    print(f"📊 Excel saved → {xlsx_path}")


# ══════════════════════════════════════════════════════════════
#  DASHBOARD OVERLAY
# ══════════════════════════════════════════════════════════════
PANEL_W = 355

def draw_dashboard(frame, fps, fc, recording, elapsed, hist, sess_frames):
    """fc = per-frame label counts dict; hist = deque of (ts, c, n, ic)"""
    fh, fw = frame.shape[:2]
    px = fw - PANEL_W

    # semi-transparent background
    ov = frame.copy()
    cv2.rectangle(ov, (px, 0), (fw, fh), D_BG, -1)
    cv2.addWeighted(ov, 0.83, frame, 0.17, 0, frame)
    cv2.line(frame, (px, 0), (px, fh), D_ACCENT, 2)

    x0 = px + 14
    y  = 28

    # Title
    cv2.putText(frame, "MASK DETECTOR", (x0, y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.60, D_ACCENT, 2, cv2.LINE_AA)
    y += 20
    cv2.putText(frame, "CROWD ANALYTICS DASHBOARD", (x0, y),
                cv2.FONT_HERSHEY_SIMPLEX, 0.38, D_TEXT, 1, cv2.LINE_AA)
    y += 14
    cv2.line(frame, (x0, y), (fw-14, y), D_ACCENT, 1); y += 14

    # REC indicator
    if recording:
        cv2.circle(frame, (x0+7, y+5), 7, (0, 0, 230), -1)
        cv2.putText(frame, f" REC   {int(elapsed//60):02d}:{int(elapsed%60):02d}.{int((elapsed%1)*10)}",
                    (x0+18, y+12), cv2.FONT_HERSHEY_SIMPLEX, 0.52, (60, 60, 255), 2, cv2.LINE_AA)
    else:
        cv2.putText(frame, "[ R ] START RECORDING", (x0, y+12),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.42, (110, 110, 120), 1, cv2.LINE_AA)
    y += 26
    cv2.line(frame, (x0, y), (fw-14, y), (45, 45, 58), 1); y += 12

    # FPS
    fps_col = (0, 220, 80) if fps >= 20 else (0, 200, 255) if fps >= 10 else (0, 0, 220)
    cv2.putText(frame, "FPS", (x0, y+13),
                cv2.FONT_HERSHEY_SIMPLEX, 0.44, D_TEXT, 1, cv2.LINE_AA)
    cv2.putText(frame, f"{fps:6.1f}", (x0+85, y+13),
                cv2.FONT_HERSHEY_SIMPLEX, 0.62, fps_col, 2, cv2.LINE_AA)
    y += 24

    # Frames recorded
    cv2.putText(frame, "FRAMES", (x0, y+13),
                cv2.FONT_HERSHEY_SIMPLEX, 0.44, D_TEXT, 1, cv2.LINE_AA)
    cv2.putText(frame, f"{sess_frames:7d}", (x0+85, y+13),
                cv2.FONT_HERSHEY_SIMPLEX, 0.62, D_TEXT, 1, cv2.LINE_AA)
    y += 24
    cv2.line(frame, (x0, y), (fw-14, y), (45, 45, 58), 1); y += 12

    # Per-label counts + bar
    total = max(sum(fc.values()), 1)
    rows  = [
        ("CORRECT",   "Correct Mask",   D_CORRECT),
        ("NO MASK",   "No Mask",         D_NOMASK),
        ("INCORRECT", "Incorrect Mask",  D_INCORRECT),
        ("UNKNOWN",   "Unknown",         D_UNKNOWN),
    ]
    bar_max = PANEL_W - 158
    for short, key, col in rows:
        cnt = fc.get(key, 0)
        cv2.putText(frame, short, (x0, y+12),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.40, col, 1, cv2.LINE_AA)
        cv2.putText(frame, str(cnt), (x0+100, y+12),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.56, col, 2, cv2.LINE_AA)
        bx = x0 + 140
        bw = int(cnt / total * bar_max)
        cv2.rectangle(frame, (bx, y+2), (bx+bar_max, y+14), (38,38,50), -1)
        if bw > 0:
            cv2.rectangle(frame, (bx, y+2), (bx+bw, y+14), col, -1)
        y += 22

    cv2.line(frame, (x0, y), (fw-14, y), (45, 45, 58), 1); y += 12

    # Rolling trend graph
    cv2.putText(frame, f"TREND  (last {GRAPH_WINDOW_SEC}s)", (x0, y+11),
                cv2.FONT_HERSHEY_SIMPLEX, 0.37, D_ACCENT, 1, cv2.LINE_AA)
    y += 16

    gx1, gy1 = x0, y
    gx2, gy2 = fw - 14, y + GRAPH_H
    cv2.rectangle(frame, (gx1, gy1), (gx2, gy2), (26, 26, 36), -1)
    cv2.rectangle(frame, (gx1, gy1), (gx2, gy2), (60, 60, 80), 1)

    # Horizontal grid lines
    for frac in [0.25, 0.5, 0.75]:
        gy = int(gy1 + GRAPH_H * (1 - frac))
        cv2.line(frame, (gx1, gy), (gx2, gy), (40, 40, 55), 1)

    if len(hist) >= 2:
        now_t    = hist[-1][0]
        win_s    = now_t - GRAPH_WINDOW_SEC
        pts      = [(t, c, n, ic) for t, c, n, ic in hist if t >= win_s]
        if len(pts) >= 2:
            maxv = max(max(c+n+ic for _, c, n, ic in pts), 1)
            gw   = gx2 - gx1
            gh   = gy2 - gy1

            def tx(t):  return gx1 + int((t - win_s) / GRAPH_WINDOW_SEC * gw)
            def vy(v):  return gy2 - int(v / maxv * (gh - 6)) - 3

            for col, idx in [(D_CORRECT, 1), (D_NOMASK, 2), (D_INCORRECT, 3)]:
                poly = [(tx(p[0]), vy(p[idx])) for p in pts]
                for i in range(len(poly)-1):
                    cv2.line(frame, poly[i], poly[i+1], col, 2, cv2.LINE_AA)

    y = gy2 + 8

    # Legend
    lx = x0
    for name, col in [("● Correct", D_CORRECT), ("● No Mask", D_NOMASK), ("● Incorrect", D_INCORRECT)]:
        cv2.putText(frame, name, (lx, y+10), cv2.FONT_HERSHEY_SIMPLEX, 0.33, col, 1, cv2.LINE_AA)
        lx += 100
    y += 20

    cv2.line(frame, (x0, y), (fw-14, y), (45, 45, 58), 1); y += 10
    cv2.putText(frame, "[R] Start/Stop Recording   [ESC] Quit", (x0, y+11),
                cv2.FONT_HERSHEY_SIMPLEX, 0.32, (90, 90, 100), 1, cv2.LINE_AA)


# ══════════════════════════════════════════════════════════════
#  CAMERA
# ══════════════════════════════════════════════════════════════
cap = cv2.VideoCapture(0)
cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)
cap.set(cv2.CAP_PROP_BUFFERSIZE,   1)

tracker     = FaceTracker()
prev_time   = 0.0
fps         = 0.0
fps_hist    = deque(maxlen=300)

# Session state
recording      = False
sess_start_ts  = None
sess_start_str = None
csv_path       = None
xlsx_path      = None
frame_idx      = 0
sess_frames    = 0
sess_counts    = {k: 0 for k in ["Correct Mask","No Mask","Incorrect Mask","Unknown"]}

# Rolling graph history: (timestamp, correct, no_mask, incorrect)
hist_counts = deque(maxlen=500)

print("━"*52)
print("  Face Mask Crowd Detector")
print("  R   → Start / Stop recording")
print("  ESC → Quit")
print("━"*52)

# ══════════════════════════════════════════════════════════════
#  MAIN LOOP
# ══════════════════════════════════════════════════════════════
while True:
    ret, frame = cap.read()
    if not ret:
        break

    frame_idx += 1
    fh, fw = frame.shape[:2]

    # YOLO detection on downscaled frame
    small   = cv2.resize(frame, (640, 360))
    results = yolo_model(small, verbose=False, conf=YOLO_CONF_THRESHOLD)
    h_r, w_r = fh/360, fw/640

    dets = []
    for r in results:
        for box in r.boxes:
            sx1, sy1, sx2, sy2 = map(int, box.xyxy[0])
            x1 = int(sx1*w_r); x2 = int(sx2*w_r)
            y1 = int(sy1*h_r); y2 = int(sy2*h_r)
            x1p, y1p, x2p, y2p = pad_bbox(x1, y1, x2, y2,
                                            FACE_PADDING_RATIO, fh, fw)
            crop = frame[y1p:y2p, x1p:x2p]
            if crop.size == 0: continue
            probs = mask_model.predict(preprocess(crop), verbose=0)[0]
            dets.append((x1, y1, x2, y2, probs))

    smoothed = tracker.update(dets)

    # FPS
    ct = time.time()
    if prev_time > 0:
        fps = 1.0 / (ct - prev_time + 1e-9)
    fps_hist.append(fps)
    prev_time = ct

    # Per-frame counts
    fc = {"Correct Mask": 0, "No Mask": 0, "Incorrect Mask": 0, "Unknown": 0}
    for *_, label, _conf in smoothed:
        fc[label] = fc.get(label, 0) + 1

    # Graph history (always, not just when recording)
    hist_counts.append((ct, fc["Correct Mask"], fc["No Mask"], fc["Incorrect Mask"]))

    # Draw bounding boxes (left part of frame only)
    for (x1, y1, x2, y2, label, conf) in smoothed:
        if x1 > fw - PANEL_W: continue
        col = BBOX_COLOR.get(label, BBOX_COLOR["Unknown"])
        cv2.rectangle(frame, (x1, y1), (x2, y2), col, 3)
        t = f"{label}  {conf*100:.1f}%" if label != "Unknown" else "Low Confidence"
        draw_label(frame, t, x1, y1, col)

    # Recording
    if recording:
        sess_frames += 1
        elapsed = ct - sess_start_ts
        for k in fc:
            sess_counts[k] += fc[k]

        total_faces = len(smoothed)
        rows = []
        ts_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]

        if total_faces == 0:
            rows.append([frame_idx, ts_str, f"{elapsed:.3f}",
                         "—","—","—", 0, 0, 0, 0, f"{fps:.1f}"])
        else:
            for fid, (_, _, _, _, label, conf) in enumerate(smoothed, 1):
                rows.append([
                    frame_idx, ts_str, f"{elapsed:.3f}",
                    fid, label, f"{conf*100:.2f}",
                    total_faces,
                    fc["Correct Mask"], fc["No Mask"], fc["Incorrect Mask"],
                    f"{fps:.1f}"
                ])
        append_csv(csv_path, rows)

    # Dashboard
    elapsed_d = (ct - sess_start_ts) if recording else 0.0
    draw_dashboard(frame, fps, fc, recording, elapsed_d, hist_counts, sess_frames)

    cv2.imshow("Face Mask Crowd Detection", frame)

    key = cv2.waitKey(1) & 0xFF

    if key == 27:  # ESC
        break

    elif key in (ord('r'), ord('R')):
        if not recording:
            # START
            recording      = True
            sess_start_ts  = time.time()
            sess_start_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sess_frames    = 0
            sess_counts    = {k: 0 for k in sess_counts}
            tag = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_path  = os.path.join(OUTPUT_DIR, f"session_{tag}.csv")
            xlsx_path = os.path.join(OUTPUT_DIR, f"session_{tag}.xlsx")
            init_csv(csv_path)
            print(f"\n🔴 Recording STARTED  →  {csv_path}")
        else:
            # STOP
            recording = False
            end_str   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            dur       = time.time() - sess_start_ts
            hh, rem   = divmod(int(dur), 3600)
            mm, ss    = divmod(rem, 60)
            summary   = {
                "duration":         f"{hh:02d}:{mm:02d}:{ss:02d}",
                "total_frames":     sess_frames,
                "total_detections": sum(sess_counts.values()),
                "correct":          sess_counts["Correct Mask"],
                "no_mask":          sess_counts["No Mask"],
                "incorrect":        sess_counts["Incorrect Mask"],
                "unknown":          sess_counts["Unknown"],
                "peak_fps":         max(fps_hist) if fps_hist else 0,
                "avg_fps":          float(np.mean(list(fps_hist))) if fps_hist else 0,
            }
            print(f"\n⏹  Recording STOPPED")
            print(f"   Duration         : {summary['duration']}")
            print(f"   Frames logged    : {summary['total_frames']}")
            print(f"   Total detections : {summary['total_detections']}")
            print(f"   ✅ Correct Mask  : {summary['correct']}")
            print(f"   ❌ No Mask       : {summary['no_mask']}")
            print(f"   ⚠  Incorrect     : {summary['incorrect']}")
            print(f"   Avg FPS          : {summary['avg_fps']:.1f}")
            print(f"📄 CSV   → {csv_path}")
            export_excel(csv_path, xlsx_path, sess_start_str, end_str, summary)
            sess_frames = 0

# ══════════════════════════════════════════════════════════════
#  AUTO-SAVE ON QUIT IF STILL RECORDING
# ══════════════════════════════════════════════════════════════
if recording:
    end_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    dur     = time.time() - sess_start_ts
    hh, rem = divmod(int(dur), 3600)
    mm, ss  = divmod(rem, 60)
    summary = {
        "duration":         f"{hh:02d}:{mm:02d}:{ss:02d}",
        "total_frames":     sess_frames,
        "total_detections": sum(sess_counts.values()),
        "correct":          sess_counts["Correct Mask"],
        "no_mask":          sess_counts["No Mask"],
        "incorrect":        sess_counts["Incorrect Mask"],
        "unknown":          sess_counts["Unknown"],
        "peak_fps":         max(fps_hist) if fps_hist else 0,
        "avg_fps":          float(np.mean(list(fps_hist))) if fps_hist else 0,
    }
    print("\n⚠  Auto-saving session before exit...")
    print(f"📄 CSV  → {csv_path}")
    export_excel(csv_path, xlsx_path, sess_start_str, end_str, summary)

cap.release()
cv2.destroyAllWindows()
